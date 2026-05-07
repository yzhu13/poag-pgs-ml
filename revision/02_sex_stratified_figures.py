"""
Revision Additions — iScience ISCIENCE-D-26-03991
Handles remaining items identified in status review:
  Analysis 8  - Sex-stratified AUC (LR + PRS616, 5x5 CV)
  Figure 2A   - Bar chart of SNP counts per chromosome (PRS616 vs PRS526)
  Figure 2B   - Improved violin plots of PGS distribution (cases/controls/suspects)
  ST3/ST4     - Numeric verification against CSV
  DIOP        - Recompute DIOP asymmetry AUC with LR as primary model

Paths (relative to this script in revision/):
  Input data : ../input date/
  Output dir : ./  (outputs land alongside this script in revision/)
  Supp tables: ../revision output/Supplemental_Tables_Revised.xlsx
"""

import warnings
warnings.filterwarnings("ignore")

from pathlib import Path
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
from scipy import stats

from sklearn.preprocessing import StandardScaler
from sklearn.impute import SimpleImputer
from sklearn.pipeline import Pipeline
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import RepeatedStratifiedKFold, cross_validate
from sklearn.utils import resample

import openpyxl

SEED = 42
np.random.seed(SEED)

# -----------------------------------------------------------------
# Paths  (relative to this script; run from the revision/ directory,
#         or adjust _HERE if needed)
# -----------------------------------------------------------------
_HERE    = Path(__file__).resolve().parent          # .../poag-pgs-ml/revision/
REPO     = _HERE.parent                             # .../poag-pgs-ml/
DATA_DIR = REPO / "input date"                      # data files
OUT_DIR  = _HERE                                    # outputs go into revision/
SUPP_DIR = REPO / "revision output"                 # supplemental tables

OUT = str(OUT_DIR) + "/"   # string path for older APIs

# -----------------------------------------------------------------
# Load data
# -----------------------------------------------------------------
print("Loading data...")
train   = pd.read_excel(DATA_DIR / "271_training_cohort_4_new_PRS_cleaned.xlsx")
suspect = pd.read_excel(DATA_DIR / "1013_testing_cohort_only_suspect_cleaned.xlsx")

PGS_MAP = {
    "POAAGG PRS": "POAAGG_PRS",
    "MEGA PRS":   "MEGA_PRS",
    "PRS526":     "PRS526",
    "PRS616":     "PRS616",
}
train = train.rename(columns=PGS_MAP)

SUS_MAP = {
    "PRS-CS POAAGG": "POAAGG_PRS",
    "PRS-CS MEGA":   "MEGA_PRS",
    "MTAG PRS":      "PRS526",
    "MEGA PRS":      "PRS616",
}
suspect = suspect.rename(columns=SUS_MAP)

DEMO  = ["Age", "Gender"]
LABEL = "CaseCtrl"
y     = train[LABEL].values


def make_lr():
    return Pipeline([
        ("imp", SimpleImputer(strategy="median")),
        ("sc",  StandardScaler()),
        ("clf", LogisticRegression(max_iter=1000, random_state=SEED)),
    ])


# =================================================================
# ANALYSIS 8 — Sex-stratified AUC  (LR + PRS616, 5x5 repeated CV)
# =================================================================
print("\n" + "="*60)
print("ANALYSIS 8: Sex-stratified AUC  (LR + PRS616)")
print("="*60)

CV55 = RepeatedStratifiedKFold(n_splits=5, n_repeats=5, random_state=SEED)

strata = {
    "Female (Gender=0)": train["Gender"] == 0,
    "Male   (Gender=1)": train["Gender"] == 1,
    "Overall":           pd.Series([True] * len(train)),
}

sex_rows = []
for label, mask in strata.items():
    sub    = train[mask]
    X_sub  = sub[DEMO + ["PRS616"]].values
    y_sub  = sub[LABEL].values
    n_case = int(y_sub.sum())
    n_ctrl = int((1 - y_sub).sum())

    if n_case < 10 or n_ctrl < 10:
        print(f"  {label}: too few samples (n_case={n_case}, n_ctrl={n_ctrl}) — skipping CV")
        sex_rows.append({"Stratum": label, "N_total": len(sub),
                         "N_cases": n_case, "N_controls": n_ctrl,
                         "AUC_mean": float("nan"), "AUC_SD": float("nan"),
                         "AUC_95CI_lo": float("nan"), "AUC_95CI_hi": float("nan")})
        continue

    scores = cross_validate(make_lr(), X_sub, y_sub,
                            cv=CV55, scoring="roc_auc", n_jobs=-1)
    aucs  = scores["test_score"]
    ci_lo, ci_hi = np.percentile(aucs, [2.5, 97.5])
    print(f"  {label}: N={len(sub)} ({n_case}C/{n_ctrl}Ctrl)  "
          f"AUC={aucs.mean():.3f}+/-{aucs.std():.3f}  95%CI [{ci_lo:.3f},{ci_hi:.3f}]")
    sex_rows.append({
        "Stratum": label.strip(), "N_total": len(sub),
        "N_cases": n_case, "N_controls": n_ctrl,
        "AUC_mean": round(aucs.mean(), 4), "AUC_SD": round(aucs.std(), 4),
        "AUC_95CI_lo": round(ci_lo, 4), "AUC_95CI_hi": round(ci_hi, 4),
    })

sex_df = pd.DataFrame(sex_rows)
sex_df.to_csv(OUT + "A8_sex_stratified_AUC.csv", index=False)
print(sex_df.to_string(index=False))
print("  -> Saved A8_sex_stratified_AUC.csv")

# Plot sex-stratified AUC
fig, ax = plt.subplots(figsize=(7, 4))
valid  = sex_df.dropna(subset=["AUC_mean"])
colors = ["#E87D4B", "#4878CF", "#5AA554"]
for i, (_, row) in enumerate(valid.iterrows()):
    ax.barh(row["Stratum"], row["AUC_mean"],
            xerr=row["AUC_SD"], color=colors[i % 3],
            edgecolor="black", linewidth=0.7, capsize=5,
            error_kw={"elinewidth": 1.2})
    ax.text(row["AUC_mean"] + row["AUC_SD"] + 0.005, i,
            f'{row["AUC_mean"]:.3f}+/-{row["AUC_SD"]:.3f}\n'
            f'(N={row["N_total"]}: {row["N_cases"]}C/{row["N_controls"]}Ctrl)',
            va="center", fontsize=8.5)
ax.axvline(0.5, color="gray", linestyle="--", linewidth=1)
ax.set_xlabel("Mean AUC (5x5 repeated CV)", fontsize=11)
ax.set_title("Sex-stratified AUC: LR + Age + PRS616", fontsize=12)
ax.set_xlim(0.45, 0.95)
plt.tight_layout()
plt.savefig(OUT + "A8_sex_stratified_AUC.png", dpi=200, bbox_inches="tight")
plt.close()
print("  -> Saved A8_sex_stratified_AUC.png")


# =================================================================
# FIGURE 2A — Bar chart: SNP counts per chromosome (PRS616 vs PRS526)
# =================================================================
print("\n" + "="*60)
print("FIGURE 2A: SNP counts per chromosome")
print("="*60)

wb = openpyxl.load_workbook(SUPP_DIR / "Supplemental_Tables_Revised.xlsx")


def read_chroms(ws):
    """Read chromosome column from a sheet, skip header rows."""
    chroms = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0] is not None:
            try:
                chroms.append(int(row[0]))
            except (ValueError, TypeError):
                pass
    return chroms


chroms_616 = read_chroms(wb["ST1A"])
chroms_526 = read_chroms(wb["ST1B"])

print(f"  PRS616: {len(chroms_616)} SNPs across {len(set(chroms_616))} chromosomes")
print(f"  PRS526: {len(chroms_526)} SNPs across {len(set(chroms_526))} chromosomes")

all_chroms = sorted(set(chroms_616) | set(chroms_526))
cnt616 = pd.Series(chroms_616).value_counts().reindex(all_chroms, fill_value=0)
cnt526 = pd.Series(chroms_526).value_counts().reindex(all_chroms, fill_value=0)

x = np.arange(len(all_chroms))
w = 0.38

fig, ax = plt.subplots(figsize=(13, 5))
bars1 = ax.bar(x - w/2, cnt616.values, w, label="PRS616 (MEGA-weighted, 616 SNPs)",
               color="#4878CF", edgecolor="white", linewidth=0.5)
bars2 = ax.bar(x + w/2, cnt526.values, w, label="PRS526 (MTAG-weighted, 526 SNPs)",
               color="#E87D4B", edgecolor="white", linewidth=0.5)

ax.set_xticks(x)
ax.set_xticklabels([str(c) for c in all_chroms], fontsize=9)
ax.set_xlabel("Chromosome", fontsize=12)
ax.set_ylabel("Number of SNPs", fontsize=12)
ax.set_title("Figure 2A: Distribution of POAG Risk SNPs by Chromosome\n"
             "PRS616 (MEGA-weighted) and PRS526 (MTAG-weighted)", fontsize=12)
ax.legend(fontsize=10)
ax.spines["top"].set_visible(False)
ax.spines["right"].set_visible(False)
ax.yaxis.grid(True, alpha=0.4, linestyle="--")
ax.set_axisbelow(True)

for bar in bars1:
    h = bar.get_height()
    if h > 0:
        ax.text(bar.get_x() + bar.get_width()/2, h + 0.2, str(int(h)),
                ha="center", va="bottom", fontsize=7, color="#4878CF")
for bar in bars2:
    h = bar.get_height()
    if h > 0:
        ax.text(bar.get_x() + bar.get_width()/2, h + 0.2, str(int(h)),
                ha="center", va="bottom", fontsize=7, color="#E87D4B")

plt.tight_layout()
plt.savefig(OUT + "Fig2A_SNP_chromosome_barchart.png", dpi=300, bbox_inches="tight")
plt.close()
print("  -> Saved Fig2A_SNP_chromosome_barchart.png")


# =================================================================
# FIGURE 2B — Improved violin plots: PGS distribution by group
# =================================================================
print("\n" + "="*60)
print("FIGURE 2B: PGS distribution violin plots")
print("="*60)

PGS_LIST   = ["POAAGG_PRS", "MEGA_PRS", "PRS526", "PRS616"]
PGS_LABELS = [
    "POAAGG PRS\n(genome-wide)",
    "MEGA PRS\n(genome-wide)",
    "PRS526\n(MTAG curated)",
    "PRS616\n(MEGA curated)",
]

# Standardize each PGS using training cohort mean/SD
for pgs in PGS_LIST:
    mu, sd = train[pgs].mean(), train[pgs].std()
    train[pgs + "_z"]   = (train[pgs] - mu) / sd
    if pgs in suspect.columns:
        suspect[pgs + "_z"] = (suspect[pgs] - mu) / sd

train["Group"] = train[LABEL].map({1: "Cases", 0: "Controls"})

fig, axes = plt.subplots(1, 4, figsize=(16, 6), sharey=False)
pal = {"Cases": "#D43F3A", "Controls": "#4878CF", "Suspects": "#5AA554"}

for ax, pgs, lbl in zip(axes, PGS_LIST, PGS_LABELS):
    col = pgs + "_z"
    df_train = train[["Group", col]].rename(columns={col: "PGS"})
    if col in suspect.columns:
        df_sus = suspect[[col]].rename(columns={col: "PGS"})
        df_sus["Group"] = "Suspects"
        df_plot = pd.concat([df_train, df_sus], ignore_index=True)
    else:
        df_plot = df_train.copy()

    order = [g for g in ["Cases", "Controls", "Suspects"] if g in df_plot["Group"].unique()]

    sns.violinplot(data=df_plot, x="Group", y="PGS", order=order,
                   palette=pal, ax=ax, inner=None, linewidth=1.2, cut=0.5, alpha=0.7)
    sns.stripplot(data=df_plot, x="Group", y="PGS", order=order,
                  palette=pal, ax=ax, size=2.2, alpha=0.45, jitter=True)

    for xi, grp in enumerate(order):
        med = df_plot.loc[df_plot["Group"] == grp, "PGS"].median()
        ax.hlines(med, xi - 0.35, xi + 0.35, colors="black", linewidths=1.5)

    if "Cases" in order and "Controls" in order:
        c = df_plot.loc[df_plot["Group"] == "Cases", "PGS"]
        k = df_plot.loc[df_plot["Group"] == "Controls", "PGS"]
        _, pval = stats.mannwhitneyu(c, k, alternative="two-sided")
        star = "***" if pval < 0.001 else ("**" if pval < 0.01 else ("*" if pval < 0.05 else "ns"))
        ymax  = df_plot["PGS"].quantile(0.97)
        xi_c  = order.index("Cases")
        xi_k  = order.index("Controls")
        ax.plot([xi_c, xi_c, xi_k, xi_k],
                [ymax, ymax + 0.1, ymax + 0.1, ymax], color="black", lw=1)
        ax.text((xi_c + xi_k) / 2, ymax + 0.12, star, ha="center", fontsize=10)

    ax.set_title(lbl, fontsize=10.5, fontweight="bold")
    ax.set_xlabel("")
    ax.set_ylabel("Standardized PGS score" if ax == axes[0] else "", fontsize=9)
    ns = [f"{g}\n(n={df_plot.loc[df_plot['Group']==g].shape[0]})" for g in order]
    ax.set_xticklabels(ns, fontsize=8.5)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)

fig.suptitle("Figure 2B: Polygenic Score Distributions by Group\n"
             "(Standardized to training cohort mean/SD; black line = median)",
             fontsize=12, y=1.01)
plt.tight_layout()
plt.savefig(OUT + "Fig2B_PGS_violins.png", dpi=300, bbox_inches="tight")
plt.close()
print("  -> Saved Fig2B_PGS_violins.png")


# =================================================================
# DIOP / DRNFL / DCDR — LR as primary model (asymmetry analysis)
# =================================================================
print("\n" + "="*60)
print("DIOP/DRNFL/DCDR: LR as primary model (asymmetry analysis)")
print("="*60)

CV1010 = RepeatedStratifiedKFold(n_splits=10, n_repeats=10, random_state=SEED)

diop_rows = []
for delta_col in ["delta_IOP", "delta_CDR", "delta_RNFL"]:
    if delta_col not in train.columns:
        print(f"  Column {delta_col} not found, skipping")
        continue
    sub   = train.dropna(subset=[delta_col])
    y_sub = sub[LABEL].values
    for feat_label, extra_cols in [("Age+Sex (baseline)", DEMO),
                                   ("Age+Sex+PRS616",     DEMO + ["PRS616"])]:
        X_sub  = sub[extra_cols + [delta_col]].values
        scores = cross_validate(make_lr(), X_sub, y_sub,
                                cv=CV1010, scoring="roc_auc", n_jobs=-1)
        aucs  = scores["test_score"]
        ci_lo, ci_hi = np.percentile(aucs, [2.5, 97.5])
        print(f"  {delta_col} | {feat_label}: N={len(sub)}  "
              f"AUC={aucs.mean():.3f}+/-{aucs.std():.3f}  95%CI [{ci_lo:.3f},{ci_hi:.3f}]")
        diop_rows.append({
            "Asymmetry_Feature": delta_col,
            "Model":    feat_label,
            "N":        len(sub),
            "AUC_mean": round(aucs.mean(), 4),
            "AUC_SD":   round(aucs.std(), 4),
            "AUC_95CI_lo": round(ci_lo, 4),
            "AUC_95CI_hi": round(ci_hi, 4),
        })

pd.DataFrame(diop_rows).to_csv(OUT + "A9_DIOP_DRNFL_DCDR_LR_AUC.csv", index=False)
print("  -> Saved A9_DIOP_DRNFL_DCDR_LR_AUC.csv")


# =================================================================
# ST3/ST4 NUMERIC VERIFICATION
# =================================================================
print("\n" + "="*60)
print("ST3/ST4 VERIFICATION: CSV vs Excel")
print("="*60)

for csv_name in ["ST3_CV_results.csv", "ST4_CV_results.csv"]:
    csv_path = SUPP_DIR / csv_name
    if csv_path.exists():
        df = pd.read_csv(csv_path)
        print(f"{csv_name}: {len(df)} rows, columns: {list(df.columns)}")
        print(df.head(3).to_string(index=False))
    else:
        print(f"{csv_name}: NOT FOUND at {csv_path}")

wb_rev     = openpyxl.load_workbook(SUPP_DIR / "Supplemental_Tables_Revised.xlsx")
sheet_names = wb_rev.sheetnames
print(f"\nExcel sheets: {sheet_names}")
for sname in ["ST3", "ST4"]:
    if sname in sheet_names:
        ws   = wb_rev[sname]
        rows = list(ws.iter_rows(values_only=True))
        print(f"  {sname} Excel: {len(rows)} total rows")
    else:
        print(f"  {sname}: NOT FOUND in Excel")

print("\nAll done.")
