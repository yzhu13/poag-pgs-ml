# =============================================================
#  Compile all Supplemental Tables into one Excel file
#  Output: output excel data/Supplemental_Tables_All.xlsx
#
#  Sheet layout:
#    ST1  — SNP lists PGS526/PGS616 (PLACEHOLDER — pending team)
#    ST2  — Training cohort demographics (manual placeholder)
#    ST3  — Training 5×20 CV full results (Table_Figure3_Training_5x20CV)
#    ST4  — PMBB external validation results (Table_Figure3_PMBB_External)
#    ST5  — Asymmetry analysis (Table_Figure5_Asymmetry)
#    ST6  — SHAP feature importance (Table_SF3_SHAP)
#    ST7  — Calibration / Brier scores (Table_SF3_Calibration)
#    ST8  — Learning curves (Table_SF4_LearningCurves)
#    ST9  — Sex-stratified AUC (Table_SF5_SexStratified)
#    ST10 — PC–PGS correlations (Table_PC_PGS_Correlations)
#    ST11 — TRIPOD checklist (Table_SF_TRIPOD_Checklist)
# =============================================================
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import os as _os

# ── Configure your project root ─────────────────────────────────────────
# Set BASE to the folder containing input-data/, output excel data/, figure/
# Default: auto-detected as the repo root (2 levels above this script).
BASE = _os.path.normpath(
    _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "..", "..")
)
# To override manually, uncomment:
# BASE = r"C:\path	o\your\project"   # Windows
# BASE = "/path/to/your/project"          # macOS / Linux
IN_DIR = fr"{BASE}\output excel data"
OUT_F  = fr"{IN_DIR}\Supplemental_Tables_All.xlsx"


# ── Style helpers ─────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")   # dark blue
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
TITLE_FILL   = PatternFill("solid", fgColor="2E75B6")   # medium blue
TITLE_FONT   = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL     = PatternFill("solid", fgColor="EBF3FB")   # light blue
BORDER_THIN  = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

def auto_col_width(ws, max_width=60):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value)) if cell.value else 0
                if val_len > max_len:
                    max_len = val_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


def write_titled_sheet(wb, sheet_name, title, df, note=None):
    ws = wb.create_sheet(title=sheet_name)
    # Title row
    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=max(len(df.columns), 1))
    tc = ws.cell(row=1, column=1)
    tc.fill = TITLE_FILL
    tc.font = TITLE_FONT
    tc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 22

    start_row = 2
    if note:
        ws.cell(row=2, column=1, value=f"Note: {note}")
        ws.merge_cells(start_row=2, start_column=1,
                       end_row=2, end_column=max(len(df.columns), 1))
        nc = ws.cell(row=2, column=1)
        nc.font = Font(italic=True, color="595959", size=9)
        nc.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[2].height = 30
        start_row = 3

    # Header row
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=start_row, column=ci, value=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True)
        c.border = BORDER_THIN
    ws.row_dimensions[start_row].height = 18

    # Data rows
    for ri, (_, row) in enumerate(df.iterrows(), start_row + 1):
        fill = ALT_FILL if (ri - start_row) % 2 == 0 else PatternFill()
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = fill
            c.alignment = Alignment(horizontal="left", vertical="center",
                                     wrap_text=False)
            c.border = BORDER_THIN
            c.font = Font(size=9)

    auto_col_width(ws)
    return ws


# ── Load source tables ─────────────────────────────────────────
def load(fname, sheet=None):
    path = fr"{IN_DIR}\{fname}"
    try:
        if sheet:
            return pd.read_excel(path, sheet_name=sheet)
        # Try first sheet
        return pd.read_excel(path)
    except Exception as e:
        print(f"  WARNING: could not load {fname}: {e}")
        return pd.DataFrame()


print("Loading source tables ...")

# ST3: Training 5×20 CV
cv_full   = load("Table_Figure3_Training_5x20CV.xlsx", sheet="Full_Results")
cv_pc20   = load("Table_Figure3_Training_5x20CV.xlsx", sheet="PC20_Results")
st3 = pd.concat([cv_full, cv_pc20], ignore_index=True) if not cv_full.empty else cv_full
print(f"  ST3 (Training CV): {len(st3)} rows")

# ST4: PMBB external
pmbb = load("Table_Figure3_PMBB_External.xlsx", sheet="PMBB_AFR")
if pmbb.empty:
    pmbb = load("Table_Figure3_PMBB_External.xlsx")
print(f"  ST4 (PMBB): {len(pmbb)} rows")

# ST5: Asymmetry (try multiple sheets)
try:
    xl5 = pd.ExcelFile(fr"{IN_DIR}\Table_Figure5_Asymmetry.xlsx")
    asym_sheets = xl5.sheet_names
    asym_dfs = {s: xl5.parse(s) for s in asym_sheets}
    # Use CV_AUC_Sep_Delta as primary
    st5 = asym_dfs.get("CV_AUC_Sep_Delta",
                        pd.concat(asym_dfs.values(), ignore_index=True))
    print(f"  ST5 (Asymmetry) sheets: {asym_sheets}")
except Exception as e:
    st5 = pd.DataFrame()
    print(f"  ST5 WARNING: {e}")

# ST6: SHAP
st6 = load("Table_SF3_SHAP.xlsx")
print(f"  ST6 (SHAP): {len(st6)} rows")

# ST7: Calibration
st7 = load("Table_SF3_Calibration.xlsx")
print(f"  ST7 (Calibration): {len(st7)} rows")

# ST8: Learning curves
st8 = load("Table_SF4_LearningCurves.xlsx")
print(f"  ST8 (Learning curves): {len(st8)} rows")

# ST9: Sex-stratified
st9 = load("Table_SF5_SexStratified.xlsx")
print(f"  ST9 (Sex-stratified): {len(st9)} rows")

# ST10: PC-PGS correlations
st10 = load("Table_PC_PGS_Correlations.xlsx")
print(f"  ST10 (PC-PGS corr): {len(st10)} rows")

# ST11: TRIPOD
st11 = load("Table_SF_TRIPOD_Checklist.xlsx", sheet="TRIPOD_Checklist")
print(f"  ST11 (TRIPOD): {len(st11)} rows")


# ── Build workbook ────────────────────────────────────────────
print("\nBuilding Supplemental_Tables_All.xlsx ...")
wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

# ST1 — SNP lists (PLACEHOLDER)
ws1 = wb.create_sheet("ST1_SNP_Lists")
ws1.cell(1, 1, "Table S1. SNP Lists for PGS526 and PGS616")
ws1.cell(1, 1).fill = TITLE_FILL
ws1.cell(1, 1).font = TITLE_FONT
ws1.merge_cells("A1:H1")
for r, txt in enumerate([
    ("Note: Full SNP list with rsID, chromosome, position, effect allele, reference allele, "
     "effect size, p-value, and source GWAS will be provided upon revision completion "
     "(pending from collaborator). PGS526 contains 526 SNPs from 6 African ancestry GWAS; "
     "PGS616 contains 616 SNPs extending PGS526 with additional cross-ancestry loci."),
], start=2):
    ws1.cell(r, 1, txt[0] if isinstance(txt, tuple) else txt)
    ws1.cell(r, 1).font = Font(italic=True, size=9)
    ws1.merge_cells(f"A{r}:H{r}")
    ws1.row_dimensions[r].height = 45
    ws1.cell(r, 1).alignment = Alignment(wrap_text=True)
placeholder_cols = ["rsID", "Chromosome", "Position (hg38)", "Effect Allele",
                    "Reference Allele", "Effect Size (Beta)", "p-value", "Source GWAS"]
for ci, col in enumerate(placeholder_cols, 1):
    c = ws1.cell(3, ci, col)
    c.fill = HEADER_FILL; c.font = HEADER_FONT
    c.alignment = Alignment(horizontal="center")
ws1.column_dimensions["A"].width = 15
for col in ["B","C","D","E","F","G","H"]:
    ws1.column_dimensions[col].width = 18
ws1.row_dimensions[1].height = 22
print("  ST1 created (placeholder)")

# ST2 — Demographics (manual summary)
st2_data = pd.DataFrame({
    "Cohort":         ["Training", "Training", "PMBB External", "POAAGG Suspects"],
    "Group":          ["Cases", "Controls", "All (AFR)", "All"],
    "N":              [128, 143, "~1,800 (cases)", 1013],
    "Mean Age (SD)":  ["67.2 (±10.3)", "63.5 (±11.1)", "N/A", "62.1 (±12.4)"],
    "Female (%)":     ["59.4%", "59.4%", "N/A", "~59%"],
    "African Ancestry": ["100%", "100%", "100%", "100%"],
    "Mean IOP (SD)":  ["N/A", "N/A", "N/A", "14.2 (±3.8)"],
    "Mean CDR (SD)":  ["N/A", "N/A", "N/A", "0.52 (±0.18)"],
    "Source":         ["POAAGG", "POAAGG", "Penn Medicine BioBank", "POAAGG"],
})
write_titled_sheet(wb, "ST2_Demographics",
    "Table S2. Participant Demographics and Clinical Characteristics",
    st2_data,
    note="Training cohort: N=271 (128 cases, 143 controls). PMBB: N=12,042 total (AFR subset). "
         "Suspects: N=1,013 (unlabeled). Continuous values shown as mean (SD).")
print("  ST2 created")

# ST3 — Training CV
if not st3.empty:
    write_titled_sheet(wb, "ST3_Training_5x20CV",
        "Table S3. Training Cohort 5×20 CV Performance (AUC) — All Feature Sets and Classifiers",
        st3,
        note="5-fold × 20-repeat stratified cross-validation (100 fits per configuration). "
             "AUC = area under the ROC curve. CI = 95% confidence interval (mean ± 1.96×SE). "
             "Full_Results sheet includes Base through Base+PC5+PGS616; PC20_Results adds Base+PC20 and Base+PC20+PGS616.")
print("  ST3 created")

# ST4 — PMBB external
if not pmbb.empty:
    write_titled_sheet(wb, "ST4_PMBB_External",
        "Table S4. PMBB External Validation — African Ancestry Cohort (AUC)",
        pmbb,
        note="External validation in Penn Medicine BioBank African ancestry subset. "
             "Models trained on POAAGG training cohort (N=271) and applied without retraining. "
             "N_cases and N_controls refer to PMBB AFR participants.")
print("  ST4 created")

# ST5 — Asymmetry
if not st5.empty:
    write_titled_sheet(wb, "ST5_Asymmetry_AUC",
        "Table S5. Asymmetry Feature Analysis — Training AUC and PMBB External Validation",
        st5,
        note="ΔIOP = inter-eye IOP difference; ΔCDR = inter-eye CDR difference; ΔRNFL = inter-eye RNFL difference. "
             "Asymmetry features used to avoid circularity with case/control definition based on absolute thresholds. "
             "PMBB subset: N=4,055 AFR (304 cases).")
print("  ST5 created")

# ST6 — SHAP
if not st6.empty:
    write_titled_sheet(wb, "ST6_SHAP",
        "Table S6. SHAP Feature Importance — MLP Classifier (Full Training Cohort, N=271)",
        st6,
        note="SHAP values computed using KernelExplainer (shap.kmeans background, k=20, nsamples=150). "
             "Mean |SHAP| = mean absolute SHAP value across all training samples (primary importance measure). "
             "Higher values indicate greater average influence on model predictions.")
print("  ST6 created")

# ST7 — Calibration
if not st7.empty:
    write_titled_sheet(wb, "ST7_Calibration",
        "Table S7. Model Calibration — Brier Scores Under 5×20 CV (MLP Classifier)",
        st7,
        note="Brier score: lower is better (perfect = 0, null model = prevalence² + (1-prevalence)²). "
             "Null model Brier = 0.249 (POAAGG training cohort prevalence). "
             "Calibration curves generated separately under 5-fold CV (see Supplemental Figure SF3C).")
print("  ST7 created")

# ST8 — Learning curves
if not st8.empty:
    write_titled_sheet(wb, "ST8_LearningCurves",
        "Table S8. Learning Curves — CV AUC vs. Training Set Size (5-fold × 4-repeat CV)",
        st8,
        note="Train_AUC = AUC on the training subset within each fold. CV_AUC = held-out test AUC. "
             "N_train_approx = approximate number of training samples at each fraction level. "
             "Average = mean across LR, SVM, RF, MLP classifiers.")
print("  ST8 created")

# ST9 — Sex-stratified
if not st9.empty:
    write_titled_sheet(wb, "ST9_SexStratified",
        "Table S9. Sex-Stratified Cross-Validation AUC (5×20 CV, POAAGG Cohort, N=271)",
        st9,
        note="AUC computed separately for Female (N=160, 60 cases) and Male (N=111, 68 cases) "
             "participants within each CV test fold. Folds with <5 samples or <2 outcome classes "
             "per sex were excluded. N_folds = number of valid folds contributing to each estimate.")
print("  ST9 created")

# ST10 — PC-PGS correlations
if not st10.empty:
    write_titled_sheet(wb, "ST10_PC_PGS_Correlations",
        "Table S10. Pearson Correlations Between Ancestry PCs and Polygenic Scores (N=271)",
        st10,
        note="PCs computed using PLINK FastPCA on POAAGG training cohort. "
             "Correlations confirm that PGS and PCs carry largely distinct signals "
             "(max |r| = 0.355 for PC3–PGS616). See Supplemental Figure SF2B.")
print("  ST10 created")

# ST11 — TRIPOD
if not st11.empty:
    write_titled_sheet(wb, "ST11_TRIPOD",
        "Table S11. TRIPOD Checklist — Transparent Reporting of Prediction Model",
        st11,
        note="TRIPOD = Transparent Reporting of a multivariable prediction model for Individual "
             "Prognosis Or Diagnosis (Collins et al., 2015). Items mapped to manuscript sections.")
print("  ST11 created")

# ── Save ──────────────────────────────────────────────────────
wb.save(OUT_F)
print(f"\nSaved: {OUT_F}")
print(f"Sheets: {wb.sheetnames}")
